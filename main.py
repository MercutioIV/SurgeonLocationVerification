# Surgeon Location Verification
# Written by Dylan Werelius on August 3rd 2022

# Imports
import requests
import json
import pandas as pd
import array as arr
import openpyxl
import time
from openpyxl.styles import PatternFill

# Starts the run timer
start = time.time()

# Opens the excel spreadsheet
workbook = pd.read_excel('C:\\Users\\dwerelius\\Desktop\\US-Vmedi-nonContacts.xlsx')
xfile = openpyxl.load_workbook('C:\\Users\\dwerelius\\Desktop\\US-Vmedi-nonContacts.xlsx')
sheet = xfile['Sheet1']
xfile['Sheet1']

# Assigns the URL (taken from NPPES API Registry)
URL = "https://npiregistry.cms.hhs.gov/api/?version=2.0"

# Initialize count variables and assign the total number of surgeons
rows = sheet.max_row - 1
unregCount = 0
movedSurgeons = 0

def getInformation():
    global unregCount
    global movedSurgeons
    for x in range(rows):
        try:
            rowNum = x+2
            npi = workbook['NPI'].iloc[x]

            # Makes the request to pull the information from the server
            queryURL = URL + f"&number={npi}"
            response = requests.get(queryURL)
            userdata = json.loads(response.text)

            # Loads the userdata from the dictionary
            firstName = (userdata["results"][0])["basic"]["first_name"]
            lastName = (userdata["results"][0])["basic"]["last_name"]
            city = (userdata["results"][0])["addresses"][0]["city"]
            state = (userdata["results"][0])["addresses"][0]["state"]
            taxDescription = (userdata["results"][0])["taxonomies"][0]["desc"]

            # Gets the location information from the excel sheet
            xcity = (workbook['Location/City'].iloc[x])
            xcity = xcity.upper()
            xstate = (workbook['State'].iloc[x])
            xstate = xstate.upper()

            # Loads the taxonomy description into the S column
            sheet[f'S{rowNum}'] = f'{taxDescription}'

            if (xcity != city):

                # Loads the correct city into the Q column and the correct state into the R column
                sheet[f'Q{rowNum}'] = f'{city}'
                sheet[f'R{rowNum}'] = f'{state}'

                # These lines are just changing the color of the cells in a row that has a incorrect address
                sheet[f'A{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'B{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'C{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'D{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'E{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'F{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'G{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'H{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'I{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'J{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'K{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'L{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'M{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'N{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'O{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'P{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'Q{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'R{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'S{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'T{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'U{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'V{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'W{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'X{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'Y{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')
                sheet[f'Z{rowNum}'].fill = PatternFill(patternType='solid',
                                                       fgColor='FC2C03')

                # Prints out the location statement
                print(f"\n{firstName} {lastName} has moved to {city}, {state}")
                movedSurgeons += 1

        except KeyError:

            # Changes the color of the city and state cells to be black
            sheet[f'D{rowNum}'].fill = PatternFill(patternType='solid',
                                                   fgColor='000000')
            sheet[f'E{rowNum}'].fill = PatternFill(patternType='solid',
                                                   fgColor='000000')

            # Prints out the inactive statement and then adds 1 to the total amount of unregistered surgeons
            print(f"\nNPI# {npi} is no longer registered")
            unregCount+=1


# Save the changes to the fil and then print out the statistics
getInformation()
xfile.save('C:\\Users\\dwerelius\\Desktop\\US-Vmedi-nonContacts.xlsx')
print("\n#############################################")
print("# Inactive Surgeons: " + str(unregCount))
print("# New Surgeon Locations: " + str(movedSurgeons))
end = time.time() # Ends the run timer
print("# Finished in " + str(end - start) + " seconds")
print("#############################################")